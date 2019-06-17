import laddersandsnakes
from laddersandsnakes import*
import checkers
from checkers import*
from openpyxl import Workbook
from openpyxl import load_workbook


#===============================functions====================
def inpt(x):
    def keyboard(x,y):
        word=''
        pygame.display.flip()
        default_font=pygame.font.get_default_font()
        font=pygame.font.SysFont(default_font,30)
        done = True
        while done:
            for event in pygame.event.get():
                if event.type==pygame.QUIT:
                    pygame.quit()
                    quit()
                if event.type == pygame.KEYDOWN:                      
                    if event.key == pygame.K_a:
                        word+=chr(event.key)
                    if event.key == pygame.K_b:
                        word+=chr(event.key)
                    if event.key == pygame.K_c:
                        word+=chr(event.key)
                    if event.key == pygame.K_d:
                        word+=chr(event.key)
                    if event.key == pygame.K_e:
                        word+=chr(event.key)
                    if event.key == pygame.K_f:
                        word+=chr(event.key)
                    if event.key == pygame.K_g:
                        word+=chr(event.key)
                    if event.key == pygame.K_h:
                        word+=chr(event.key)
                    if event.key == pygame.K_i:
                        word+=chr(event.key)
                    if event.key == pygame.K_m:
                        word+=chr(event.key)
                    if event.key == pygame.K_l:
                        word+=chr(event.key)
                    if event.key == pygame.K_n:
                        word+=chr(event.key)
                    if event.key == pygame.K_o:
                        word+=chr(event.key)
                    if event.key == pygame.K_p:
                        word+=chr(event.key)
                    if event.key == pygame.K_q:
                        word+=chr(event.key)
                    if event.key == pygame.K_r:
                        word+=chr(event.key)
                    if event.key == pygame.K_s:
                        word+=chr(event.key)
                    if event.key == pygame.K_t:
                        word+=chr(event.key)
                    if event.key == pygame.K_w:
                        word+=chr(event.key)
                    if event.key == pygame.K_x:
                        word+=chr(event.key)
                    if event.key == pygame.K_y:
                        word+=chr(event.key)
                    if event.key == pygame.K_z:
                        word+=chr(event.key)
                    if event.key == pygame.K_0:
                        word+=chr(event.key)
                    if event.key == pygame.K_1:
                        word+=chr(event.key)
                    if event.key == pygame.K_2:
                        word+=chr(event.key)
                    if event.key == pygame.K_3:
                        word+=chr(event.key)
                    if event.key == pygame.K_4:
                        word+=chr(event.key)
                    if event.key == pygame.K_5:
                        word+=chr(event.key)
                    if event.key == pygame.K_6:
                        word+=chr(event.key)
                    if event.key == pygame.K_7:
                        word+=chr(event.key)
                    if event.key == pygame.K_8:
                        word+=chr(event.key)
                    if event.key == pygame.K_9:
                        word+=chr(event.key)
                    if event.key == pygame.K_KP_ENTER:
                        done = False         
                    if event.key == pygame.K_SPACE:
                        word+=chr(event.key)
                    if event.key == pygame.K_BACKSPACE:
                        word=""
                        pygame.draw.rect(gameDisplay, colors['gray_l'],(250,x,290,60))
                    if event.key == pygame.K_RETURN:
                        done = False                   
                    gameDisplay.blit(font.render(word,True,colors['black']),(270,y))
                    pygame.display.flip()
        return word
    word=""
    LastName=""
    pygame.draw.rect(gameDisplay, colors['gray_l'],(250,180,290,60))
    pygame.draw.rect(gameDisplay, colors['gray_l'],(250,250,290,60))
    if x==1:
     text("Player 1 enter your name :",270,200) #example asking name of player1
    if x==2:
     text("Player 2 enter your name:",270,200) #example asking name of player2
    if x==3:
     text("Enter the Password:",285,200) #example asking password
    if x==4:
     text("Enter your Kid name:",280,200) 
    if(x!=3):
            pygame.draw.rect(gameDisplay, colors['gray_l'],(250,320,290,60))
            word=keyboard(250,270)
            LastName=keyboard(320,340)
            return word,LastName
    else:
        return keyboard(250,270)  
 
#========= choosing game ===========
def Select_a_game(level):
    Game_m = True
    p_1,p_2,Q_1,Q_2=0,0,0,0
    chose1=0
    chose2=0
    chose3=0
    chose4=0
    while Game_m:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                QuitTheGame()
        num=0        
        gameDisplay.fill(colors['white'])
        gameDisplay.blit(Selectagame,(0,0))
        chose1= button_2("Ladders and snakes",500,410,200,60,colors['gray_l'],colors['gray'],'1')
        chose2= button_2("Checkers",130,410,125,60,colors['gray_l'],colors['gray'],'2')
        #chose3= button_2("Back",340,230,125,60,colors['gray_l'],colors['gray'],'3')
        chose4= button_2("Exit",340,300,125,60,colors['red'],colors['red_l'],'4')
        print(num)
        if 1==chose1:
             level,p_1,p_2,Q_1,Q_2=game_loop(level,main_menu)
             #p_1,p_2,Q_1,Q_2=8,9,10,10
             return level,p_1,p_2,Q_1,Q_2,1
        if 2==chose2:
             level,Q_1,Q_2,p_1,p_2=rival(level,main_menu)
             return level,p_1,p_2,Q_1,Q_2,2
        if 3==chose3:
             return level,p_1,p_2,Q_1,Q_2
        if 4==chose4:
            QuitTheGame()
        pygame.display.update()
        clock.tick(15)




                        
#================================               
#============save data====    


    
def read_from_xl(file_name, sheet_number = 0):
    workbook = load_workbook(file_name)
    sheet = workbook.get_sheet_names()[sheet_number]
    worksheet = workbook.get_sheet_by_name(sheet)
    rows = []
    for row in worksheet.iter_rows():
        cells = []
        for cell in row:
            cells.append(cell.value)
        rows.append(cells)
    workbook.save(file_name)
    return rows

#a = read_from_xl('test.xlsx')
#for i in range(len(a)):
#    if a[i][0] == 'shimon' and a[i][1] == 'desta':
#        print('found')


def write_to_xl(file_name,lst):
    wb = Workbook()
    ws = wb.active
    for i in lst:
         ws.append(i) 
    wb.save(file_name)

def find_kid(name,lastname):
    datalist=read_from_xl('Date.xlsx')
    KidList=[]
    for i in range(len(datalist)):
        if datalist[i][0] == name and datalist[i][1] == lastname:
            KidList.append(datalist[i])
          
    return KidList

def avg_per_age(age,game):
    sum,count=0,0
    datalist=read_from_xl('Date.xlsx')
    for i in range(len(datalist)):
         if datalist[i][4]==age and datalist[i][5]==game :
             print(grade(datalist[i]))
             sum+=grade(datalist[i])
             count+=1
    if count!=0:
        return round(sum/count,2)
    else:
        return 0
def grade(KidList):
    if KidList[3]!=0:
        return round((KidList[2]/KidList[3]),2)*100
    else: return 0

'''
def grade(name,lastname):
    KidList=find_kid(name,lastname)
    grade=[]
    KidList[i][6]
    for i in range(len(KidList)):
        grade.append((grade datalist[i][2]*grade datalist[i][3])*100)
def avarage():
    datalist=read_from_xl('Date.xlsx')
    for i in range(len(datalist)):
        sum+=
'''
#===================menu===========:============
def Kid_():
    kid1=[]
    kid2=[]
    Game_m = True
    gameDisplay.fill(colors['white'])
    message_display("Hello Kids")
    Name_1,LastName_1=inpt(1)
    Name_2,LastName_2=inpt(2)
    datalist=read_from_xl('Date.xlsx')
    p_1,p_2=0,0
    Q_1,Q_2=0,0
    level=1
    type_game=0
    games={1:'Ladders and snakes',2:'Checkers'}
    gameDisplay.fill(colors['white'])
    while Game_m:
        
        
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                QuitTheGame()
                
        gameDisplay.fill(colors['white'])
        largeText = pygame.font.Font('freesansbold.ttf',80)
        TextSurf, TextRect = text_objects("Select Level", largeText)
        TextRect.center = ((display_width/2),(display_height/2-200))
        gameDisplay.blit(TextSurf, TextRect)
        mouse = pygame.mouse.get_pos()
      
        
        #==(child menu )====

        if 1==button_2("age 6-8",340,230,125,60,colors['gray_l'],colors['gray'],'1'):   
             level,p_1,p_2,Q_1,Q_2,type_game=Select_a_game(1)
             kid1.extend([Name_1,LastName_1,p_1,Q_1,'age 6-8',games[type_game]])
             kid2.extend([Name_2,LastName_2,p_2,Q_2,'age 6-8',games[type_game]])
             datalist.append(kid1)
             datalist.append(kid2)
             write_to_xl('Date.xlsx',datalist)
             main_menu()

        if 2==button_2("age 8-10",340,300,125,60,colors['gray_l'],colors['gray'],'2'):
             level,p_1,p_2,Q_1,Q_2,type_game=Select_a_game(2)
             kid1.extend([Name_1,LastName_1,p_1,Q_1,'age 8-10',games[type_game]])
             kid2.extend([Name_2,LastName_2,p_2,Q_2,'age 8-10',games[type_game]])
             datalist.append(kid1)
             datalist.append(kid2)
             write_to_xl('Date.xlsx',datalist)
             main_menu()

        if 3==button_2("age 10-12",340,370,125,60,colors['gray_l'],colors['gray'],'3'):
             level,p_1,p_2,Q_1,Q_2,type_game=Select_a_game(3)
             kid1.extend([Name_1,LastName_1,p_1,Q_1,'age 10-12',games[type_game]])
             kid2.extend([Name_2,LastName_2,p_2,Q_2,'age 10-12',games[type_game]])
             datalist.append(kid1)
             datalist.append(kid2)
             write_to_xl('Date.xlsx',datalist)
             main_menu()

        if 4==button_2("Back",340,440,125,60,colors['red'],colors['red_l'],'4'):   
             main_menu()      
        pygame.display.update()
        clock.tick(15)
        
    
def Parent_():
    Game_m = True
    KidList=None
    gameDisplay.fill(colors['white'])
    message_display("Enter the Password")
    password=inpt(3)
    if password == '123456' :
        gameDisplay.fill(colors['white'])
        message_display("Hello Parents")
        Name_1,LastName_1=inpt(4)
        gameDisplay.fill(colors['white'])     
        KidList=find_kid(Name_1,LastName_1)
        
        if Name_1 in(KidList) and LastName_1 in(KidList):
                gameDisplay.blit(data_P,(100,200))#table
                while Game_m:
                    for event in pygame.event.get():
                        if event.type == pygame.QUIT:
                            QuitTheGame()
                            
                    gameDisplay.fill(colors['white'])
                    gameDisplay.blit(dataP,(0,0))
                    largeText = pygame.font.SysFont("comicsansms",60)
                    TextSurf, TextRect = text_objects("Your kids results:", largeText)
                    TextRect.center = ((display_width/2),(display_height/2-250))
                    gameDisplay.blit(TextSurf, TextRect)
                    
                    
                    button(str(KidList[0][0]),50,120,150,60,colors['gray_l'],colors['gray'])
                    button(str(KidList[0][1]),200,120,150,60,colors['gray_l'],colors['gray'])
                    
                    for i in range(len(KidList)):
                     Message_(str(i+1),640,220+(i*20))#screem  
                     Message_(str(KidList[i][5]),480,220+(i*20))  #game name      
                     Message_((str(KidList[i][2])+'/'+str(KidList[i][3])),300,220+(i*20))#number of questions
                     Message_(str(KidList[i][2]*2),160,220+(i*20))#score
                     Message_(str(round((KidList[i][2]/KidList[i][3]),2)*100),30,220+(i*20))#grade
                     
                       

                    button("Back",100,500,125,60,colors['gray_l'],colors['gray'],main_menu)
                    button("Exit",595,500,125,60,colors['gray_l'],colors['gray'],QuitTheGame)
                   
                    pygame.display.update()
                    clock.tick(15)
        else:
                message_display("WRONG Name Try Again")
                time.sleep(1)
    else:        
        message_display("WRONG PASSWORD!")
        time.sleep(1)        
def Counselor_():
    ages=('age 6-8','age 8-10','age 10-12')
    games=('Ladders and snakes','Checkers')
    Game_m = True
    gameDisplay.fill(colors['white'])
    message_display("Hello Counselor")
    Pass=inpt(3)
    gameDisplay.fill(colors['white'])
    #===)verify password )====
    if (Pass=="12345"):
            while Game_m:
                for event in pygame.event.get():
                    if event.type == pygame.QUIT:
                        QuitTheGame()
                gameDisplay.fill(colors['white'])
                gameDisplay.blit(dataC,(0,0))
                largeText = pygame.font.SysFont("comicsansms",60)
                TextSurf, TextRect = text_objects("The kids results:", largeText)
                TextRect.center = ((display_width/2),(display_height/2-250))
                gameDisplay.blit(TextSurf, TextRect)

                for i in range(3):
                     Message_(str(i+1),640,240+(i*20))#מסד  
                     Message_(str(ages[i]),520,240+(i*20))       
                     Message_(str(avg_per_age(ages[i],games[0])),310,240+(i*20))
                     Message_(str(avg_per_age(ages[i],games[1])),90,240+(i*20))
                #Message_(round(((avg_per_age(ages[0],games[0])+avg_per_age(ages[1],games[0])+avg_per_age(ages[2],games[0]))/3),2),400,500)
                #Message_(round(((avg_per_age(ages[0],games[1])+avg_per_age(ages[1],games[1])+avg_per_age(ages[2],games[1]))/3),2),400,560)
                button("Back",130,500,125,60,colors['gray_l'],colors['gray'],main_menu)
                button("Exit",580,500,125,60,colors['gray_l'],colors['gray'],QuitTheGame)
               
                pygame.display.update()
                clock.tick(15)
    else:
            message_display("WRONG PASSWORD!")
            time.sleep(1)
            
            
#==============================================

                
def main_menu():


    gameDisplay.fill(colors['white'])
    gameDisplay.blit(Logo,(0,0))
    pygame.display.update()
    time.sleep(4)
    
    
    
    
    Game_m = True
    while Game_m:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                QuitTheGame()
                
        gameDisplay.fill(colors['white'])
        largeText = pygame.font.Font('freesansbold.ttf',80)
        TextSurf, TextRect = text_objects("welcome", largeText)
        TextRect.center = ((display_width/2),(display_height/2-200))
        gameDisplay.blit(TextSurf, TextRect)
        mouse = pygame.mouse.get_pos()
        
        button("kid",340,230,125,60,colors['gray_l'],colors['gray'],Kid_)
        button("Parent",340,300,125,60,colors['gray_l'],colors['gray'],Parent_)                     
        button("Counselor",340,370,125,60,colors['gray_l'],colors['gray'],Counselor_)
        button("Exit",340,440,125,60,colors['red'],colors['red_l'],QuitTheGame)


        pygame.display.update()
        clock.tick(15)
        
main_menu()
#game_loop()
pygame.quit()
quit()
