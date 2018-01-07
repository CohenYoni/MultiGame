######################## import python libraries ########################
import pygame, random, time, pickle, array
from openpyxl import Workbook
from openpyxl import load_workbook

######################## dimensions ########################
display_width = 800
display_height = 600

######################## colors ########################
colors = {'black':(0,0,0), 'white':(255,255,255), 'red':(150,0,0), 'red_l':(255,70,70), 'dodger_blue':(28,134,238),
          'gray':(102,102,102), 'gray_l':(204,204,204), 'blue':(0,0,255), 'green':(0,255,0)}

######################## images ########################
exitImg = pygame.image.load('Pic_menu\Exit.png')
iconImg = pygame.image.load('Pic_menu\SCEiconImg.png')
logoImg = pygame.image.load('Pic_menu\Logo2.png')
EndImg = pygame.image.load('Pic_menu\BackgroundEnd.png')
Background_menusImg = pygame.image.load('Pic_menu\BackgroundMain.png')
helpImg = pygame.image.load('Pic_menu\help.png')
menuButtonImg = pygame.image.load('Pic_menu\menu.png')
Checkers_instructionImg = pygame.image.load('Pic_menu\checkers_instruction.png')
Ledders_instructionImg = pygame.image.load('Pic_menu\laddersAndSnakes_instruction.png')
left = 1 # left click of the mouse

######################## GUI data ########################
pygame.init() # open pygame
gameDisplay = pygame.display.set_mode((display_width, display_height)) # set a screen
pygame.display.set_caption('SCE Group 4 Multi Game')
pygame.display.set_icon(iconImg)
clock = pygame.time.Clock()
gameDisplay.fill(colors['white'])
pygame.display.update()
default_font_type = 'DavidLibre.ttf'
button_font_size = 40

# option B: use pygame.font.get_default_font()

######################## functions ########################

def text_objects(message, font_size, color = colors['black'], font_type = default_font_type):
    ''' create a text object '''
    font = pygame.font.Font(font_type, font_size)
    textSurfrace = font.render(message, True, color)
    return textSurfrace, textSurfrace.get_rect()

def text_display(message, x, y, font_size = 30, color = colors['black'], font_type = default_font_type):
    ''' show text on the screen '''
    TextSurf, TextRect = text_objects(message, font_size, color, font_type)
    TextRect.left, TextRect.top = x, y
    #pygame.draw.rect(gameDisplay, colors['white'], TextRect)
    gameDisplay.blit(TextSurf, TextRect)

def text_center(message, x, y, font_size = 30, color = colors['black'], font_type = default_font_type):
    ''' show centered text on the screen '''
    TextSurf, TextRect = text_objects(message, font_size, color, font_type)
    TextRect.center = x, y
    #pygame.draw.rect(gameDisplay, colors['white'], TextRect)
    gameDisplay.blit(TextSurf, TextRect)

def button(msg, x, y, width, height, pressed, regular_color, active_color, action = None, *action_args):
    ''' active button. pressed is a boolean variable, if it True- the button was press.'''

    mouse = pygame.mouse.get_pos()
    if x < mouse[0] < x + width and y < mouse[1] < y + height: # if mouse position is on the button
        pygame.draw.rect(gameDisplay, active_color, (x, y, width, height))
        if pressed:
            if action != None:
                action(*action_args)
            else: return True
    else:
        pygame.draw.rect(gameDisplay, regular_color, (x, y, width, height))

    # show text on the button
    textSurf, textRect = text_objects(msg, button_font_size)
    textRect.center = ((x + (width / 2)), (y + (height / 2)))
    gameDisplay.blit(textSurf, textRect)

def image_button(image, x, y, pressed, action = None, *action_args):
    ''' active image button. pressed is a boolean variable, if it True- the button was press.'''

    gameDisplay.blit(image, (x, y))
    mouse = pygame.mouse.get_pos()
    if x < mouse[0] < x + image.get_width() and y < mouse[1] < y + image.get_height() and pressed: # if mouse position is on the image
        if action != None:
            action(*action_args)
        else: return True

def QuitTheGame():
    ''' exit screen and close the game program '''
    write_to_xl('DB.xlsx', data_base)
    gameDisplay.blit(exitImg, (0, 0))
    pygame.display.update()
    time.sleep(2)
    gameDisplay.blit(EndImg, (0, 0))
    pygame.display.update()
    time.sleep(2)
    pygame.quit()
    quit()

def input_keybord(x, y):
    ''' get input from the user '''
    len_str = 15
    font = pygame.font.Font(default_font_type, 32)
    input_box = pygame.Rect(x, y, 300, 50)
    text = ''
    done = False
    while not done:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                QuitTheGame()
            if event.type == pygame.KEYDOWN:
                if event.key == pygame.K_RETURN and text != '':
                    return text
                elif event.key == pygame.K_BACKSPACE:
                    text = text[:-1]
                elif len(text) <= len_str:
                    ch = event.unicode
                    if 'A' <= ch <= 'Z' or 'a' <= ch <= 'z' or '0' <= ch <= '9': # check if the key was a correct character
                        text += ch
        # Render the current text.
        txt_surface = font.render(text, True, colors['black'])
        # Blit the input_box rect.
        pygame.draw.rect(gameDisplay, colors['white'], input_box)
        pygame.draw.rect(gameDisplay, colors['black'], input_box, 2)
        # Blit the text.
        gameDisplay.blit(txt_surface, (input_box.x+5, input_box.y+5))
        pygame.display.flip()
        clock.tick(60)

def instruction(image):
    ''' show instruction '''
    end_loop = False
    pressed = False
    while not end_loop:
        for event in pygame.event.get():
            if event.type == pygame.QUIT:
                end_loop = True
            elif event.type == pygame.MOUSEBUTTONDOWN:
                pressed = True
            elif event.type == pygame.MOUSEBUTTONUP:
                pressed = False
        gameDisplay.blit(image, (0, 0))
        if image_button(menuButtonImg, 30, 30, pressed):
            end_loop = True
            pressed = False

        pygame.display.update()
        clock.tick(60)

    return

######################## database ########################
def read_from_xl(file_name, sheet_number = 0):
    ''' load the data from the exel '''
    try: #try to open the exel
        workbook = load_workbook(file_name)
    except: #create a new file if there is no file in the folder.
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "GamesData"
        workbook.save(file_name)
    try:
        sheet = workbook.get_sheet_names()[sheet_number]
        worksheet = workbook.get_sheet_by_name(sheet)
        rows = []
        for row in worksheet.iter_rows():
            cells = []
            for cell in row:
                cells.append(cell.value)
            rows.append(cells)
        workbook.close()
        return rows # row is a list of all the row in the file, each row is also list, contain all the cell in the row
    except: # if was a problem in the file.
        gameDisplay.fill(colors['white'])
        text_display('Error. Can not open the database.', 5, 5)
        pygame.display.update()
        time.sleep(2)
        pygame.quit()
        quit()

def write_to_xl(file_name, list):
    ''' write all th data to exel. the new file overwrites the existing file.  '''
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GamesData"
    for row in list:
        sheet.append(row)
    workbook.save(file_name)

def add_game_data_to_DB(first_name, last_name, age, num_of_questions, num_of_correct_ans, game_type, quastions_type):
    data_base.append([first_name, last_name, age, num_of_questions, num_of_correct_ans, game_type, quastions_type])

def avg_per_age(age, game_type, DB):
    ''' calculate the average of all games in given age '''
    sum = count = 0
    for row in DB:
         if row != [] and row[2] == ('age ' + age) and row[5] == game_type:
             sum += grade_one_game(row)
             count += 1
    if count != 0: return round(sum / count, 2)
    else: return 0

def avg_per_game(game_type, DB):
    ''' calculate the average of all games in given game type '''
    sum = count = 0
    for row in DB:
         if row != [] and row[5] == game_type:
             sum += grade_one_game(row)
             count += 1
    if count != 0: return round(sum / count, 2)
    else: return 0

def grade_one_game(kid_data_list):
    ''' calculate the grade of one given game '''
    if kid_data_list != [] and kid_data_list[3] != 0: # if were questions
        return round((kid_data_list[4] / kid_data_list[3]) ,2) * 100
    else: return 0

def find_kid_games(first_name, last_name, DB):
    ''' return list of all the game of given child '''
    kid_data_game = []
    for row in DB:
        if row != [] and row[0] == first_name and row[1] == last_name:
            kid_data_game.append(row)
    return kid_data_game

data_base = read_from_xl('DB.xlsx') #list of all the data

######################## intro ########################
gameDisplay.blit(logoImg, (0, 0))
pygame.display.update()
time.sleep(2)