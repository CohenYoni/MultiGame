from openpyxl import load_workbook, Workbook
def read_from_xl(file_name, sheet_number = 0):
    workbook = load_workbook(file_name)
    sheet = workbook.get_sheet_names()[sheet_number]
    worksheet = workbook.get_sheet_by_name(sheet)
    rows = []
    for row in worksheet.iter_rows():
        cells = []
        for cell in row:
            cells.append(cell.value)
        rows.append(cells)
    workbook.save(file_name)
    return rows
a = read_from_xl('test.xlsx')
for i in range(len(a)):
    if a[i][0] == 'shimon' and a[i][1] == 'desta':
        print('found')
